"""Parse the MDE approved-course spreadsheets into approved_lists.yaml.

Three lists are referenced by the elective policy but not contained in it:

    MDE Approved Project-Based Electives.xlsx
        "GSD Project-Based"            -> project_based   (policy rule 1a)

    MDE Approved SEAS Courses.xlsx
        "0-1000 Level Electives"       -> seas_0_100      (policy rule 2b-i)
        "Graduate Technical Courses"   -> technical       (policy rule 2a-ii/2b-ii)

`technical` is the union of BOTH SEAS sheets: the 0-1000 sheet carries its own
"Technical course fulfillment" Yes/No column, so most 100-level courses also
satisfy the technical requirement.

GSD entries are listed by catalog NUMBER only ("6317"), with the subject implied.
The subject prefix is deliberately NOT reconstructed: two thirds of the list is
not offered in any single term, so it cannot be verified against the catalog, and
inventing "SCI6317" would risk a wrong match. They are stored as numbers and
matched on school=GSD + catalog instead.

    python -m ingest.approved              # rewrite approved_lists.yaml
    python -m ingest.approved --check      # parse and report, write nothing
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import yaml

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DEFAULT_TERM, ROOT
from ingest.db import connect

PROJECT_XLSX = ROOT / "MDE Approved Project-Based Electives.xlsx"
SEAS_XLSX = ROOT / "MDE Approved SEAS Courses.xlsx"
OUT_PATH = ROOT / "approved_lists.yaml"


def norm_code(s: str) -> str:
    """Must match api.policy.norm_code exactly."""
    return re.sub(r"[\s\-_.]", "", (s or "")).upper()


def expand_code(raw: str) -> list[str]:
    """Split the shorthand the spreadsheets use into individual course codes.

    "APCOMP 209 A and B"  -> ["APCOMP209A", "APCOMP209B"]
    "COMPSCI 1090A"       -> ["COMPSCI1090A"]
    """
    raw = (raw or "").strip()
    if not raw:
        return []
    m = re.match(r"^(.*?\d+)\s+([A-Z])\s+and\s+([A-Z])$", raw, re.I)
    if m:
        base = norm_code(m.group(1))
        return [base + m.group(2).upper(), base + m.group(3).upper()]
    return [norm_code(raw)]


def expand_gsd_number(raw: str) -> list[str]:
    """GSD numbers, minus module markers.

    "2121 [M1]"     -> ["2121"]      ([M1]/[M2] are half-term modules)
    "2227 / 2224"   -> ["2227", "2224"]
    """
    raw = (raw or "").strip()
    raw = re.sub(r"\[M\d\]", "", raw)
    return re.findall(r"\d+", raw)


def parse_project_based(path: Path = PROJECT_XLSX) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["GSD Project-Based"]
    numbers: list[str] = []
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code_cell, _type, dept, title, _instr, _units, is_pb, notes = (list(r) + [None] * 8)[:8]
        if not code_cell or str(code_cell).strip() == "Course#":
            continue
        # Defensive: honour the column rather than assuming every row is a Yes.
        if str(is_pb or "").strip().lower() not in ("yes", "y"):
            continue
        for n in expand_gsd_number(str(code_cell)):
            if n not in numbers:
                numbers.append(n)
            rows.append({"number": n, "dept": str(dept or "").strip(),
                         "title": str(title or "").strip(),
                         "notes": str(notes or "").strip()})
    return numbers, rows


def parse_seas(path: Path = SEAS_XLSX) -> tuple[list[str], list[str], list[dict]]:
    """Returns (seas_0_100 codes, technical codes, detail rows)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    low: list[str] = []
    tech: list[str] = []
    rows: list[dict] = []

    for sheet, is_low in (("0-1000 Level Electives", True),
                          ("Graduate Technical Courses", False)):
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=4, values_only=True):
            term, tech_flag, category, code_cell, title, notes = (list(r) + [None] * 6)[:6]
            if not code_cell:
                continue
            codes = expand_code(str(code_cell))
            if not codes:
                continue
            is_tech = str(tech_flag or "").strip().lower() in ("yes", "y")
            for c in codes:
                if is_low and c not in low:
                    low.append(c)
                if is_tech and c not in tech:
                    tech.append(c)
                rows.append({"code": c, "sheet": sheet, "technical": is_tech,
                             "category": str(category or "").strip(),
                             "title": str(title or "").strip(),
                             "usual_term": str(term or "").strip(),
                             "notes": str(notes or "").strip()})
    return low, tech, rows


def validate(conn, codes: list[str], gsd_numbers: list[str], term: str) -> dict:
    """Which entries can be matched against the catalog for a given term?"""
    have_codes = {norm_code(r[0]) for r in conn.execute(
        "SELECT DISTINCT code FROM courses WHERE term = ?", (term,))}
    have_gsd = {str(r[0]) for r in conn.execute(
        "SELECT DISTINCT catalog FROM courses WHERE term = ? AND school = 'GSD'", (term,))}
    return {
        "codes_matched": sorted(c for c in codes if c in have_codes),
        "codes_unmatched": sorted(c for c in codes if c not in have_codes),
        "gsd_matched": sorted(n for n in gsd_numbers if n in have_gsd),
        "gsd_unmatched": sorted(n for n in gsd_numbers if n not in have_gsd),
    }


HEADER = """\
# Approved course lists referenced by the MDE Elective Policy.
# =============================================================================
# GENERATED FILE -- do not hand-edit. Regenerate with:
#
#     python -m ingest.approved
#
# Source workbooks:
#   MDE Approved Project-Based Electives.xlsx  -> project_based
#   MDE Approved SEAS Courses.xlsx             -> seas_0_100, technical
#
# `technical` is the union of BOTH SEAS sheets: the 0-1000 sheet has its own
# "Technical course fulfillment" column, so most 100-level courses count too.
#
# GSD project-based courses are listed by catalog NUMBER, not full code. The
# spreadsheet gives only the number ("6317") and the subject prefix is not
# reconstructed -- most of the list is not offered in any single term, so a
# guessed prefix could not be verified. They match on school=GSD + catalog.
# =============================================================================
"""


def build(term: str = DEFAULT_TERM, write: bool = True) -> dict:
    pb_numbers, pb_rows = parse_project_based()
    low, tech, seas_rows = parse_seas()

    conn = connect()
    v_pb = validate(conn, [], pb_numbers, term)
    v_low = validate(conn, low, [], term)
    v_tech = validate(conn, tech, [], term)
    conn.close()

    data = {
        "project_based": {
            "verified": True,
            "source": PROJECT_XLSX.name,
            "description": "Rule 1a -- one of the two GSD courses must be "
                           "project-based, for students without a physical design background.",
            "gsd_catalog_numbers": pb_numbers,
            "codes": [],
        },
        "technical": {
            "verified": True,
            "source": f"{SEAS_XLSX.name} (both sheets)",
            "description": "Rule 2a-ii / 2b-ii -- at least one SEAS course must "
                           "be technical. Applies to every student.",
            "codes": tech,
        },
        "seas_0_100": {
            "verified": True,
            "source": f"{SEAS_XLSX.name} -- '0-1000 Level Electives'",
            "description": "Rule 2b-i -- students WITHOUT a SEAS background may "
                           "count an approved 0-100-level SEAS course.",
            "codes": low,
        },
    }

    if write:
        body = yaml.safe_dump(data, sort_keys=False, default_flow_style=False, width=100)
        OUT_PATH.write_text(HEADER + "\n" + body)

    print(f"project_based : {len(pb_numbers)} GSD catalog numbers")
    print(f"                {len(v_pb['gsd_matched'])} offered in {term}, "
          f"{len(v_pb['gsd_unmatched'])} not offered this term")
    print(f"technical     : {len(tech)} codes  "
          f"({len(v_tech['codes_matched'])} in {term}, {len(v_tech['codes_unmatched'])} not)")
    print(f"seas_0_100    : {len(low)} codes  "
          f"({len(v_low['codes_matched'])} in {term}, {len(v_low['codes_unmatched'])} not)")
    if write:
        print(f"\nwrote {OUT_PATH}")

    return {"project_based": v_pb, "technical": v_tech, "seas_0_100": v_low,
            "counts": {"project_based": len(pb_numbers), "technical": len(tech),
                       "seas_0_100": len(low)}}


def main() -> None:
    ap = argparse.ArgumentParser(description="Parse MDE approved-course spreadsheets")
    ap.add_argument("--term", default=DEFAULT_TERM)
    ap.add_argument("--check", action="store_true", help="parse and report, write nothing")
    ap.add_argument("--verbose", action="store_true", help="list unmatched entries")
    args = ap.parse_args()
    res = build(args.term, write=not args.check)
    if args.verbose:
        print("\nnot offered in", args.term)
        print("  GSD numbers :", ", ".join(res["project_based"]["gsd_unmatched"]) or "none")
        print("  technical   :", ", ".join(res["technical"]["codes_unmatched"]) or "none")
        print("  seas_0_100  :", ", ".join(res["seas_0_100"]["codes_unmatched"]) or "none")


if __name__ == "__main__":
    main()
